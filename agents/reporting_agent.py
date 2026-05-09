"""
Stage 5 — Reporting Agent

Write two formatted .xlsx reports to data/output/:
  • billing_report_YYYYMMDD.xlsx
  • reconciliation_report_YYYYMMDD.xlsx

Both have bold headers and auto-fit column widths via openpyxl.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from logger import get_logger
from dotenv import load_dotenv

load_dotenv()
log = get_logger(__name__)

OUTPUT_FOLDER = os.environ.get("OUTPUT_FOLDER", "data/output")


def _write_formatted_xlsx(df: pd.DataFrame, filepath: str) -> None:
    """Write a DataFrame to an xlsx with bold headers and auto column widths."""
    wb = Workbook()
    ws = wb.active

    # Write header row (bold)
    bold = Font(bold=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    wb.save(filepath)
    log.info("Wrote report: %s", filepath)


def run(
    df: pd.DataFrame,
    billing_summary: dict,
    recon_summary: dict,
) -> tuple[str, str]:
    """Generate billing and reconciliation Excel reports."""
    log.info("Reporting started")
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    datestamp = datetime.today().strftime("%Y%m%d")

    # --- Billing report ---
    billing_path = os.path.join(OUTPUT_FOLDER, f"billing_report_{datestamp}.xlsx")
    _write_formatted_xlsx(df, billing_path)

    # --- Reconciliation report ---
    recon_path = os.path.join(OUTPUT_FOLDER, f"reconciliation_report_{datestamp}.xlsx")

    deleted = recon_summary.get("deleted", [])
    missing = recon_summary.get("missing", [])

    # Build a combined DataFrame for the recon report
    max_len = max(len(deleted), len(missing), 1)
    recon_df = pd.DataFrame(
        {
            "deleted_card_id": deleted + [""] * (max_len - len(deleted)),
            "missing_card_id": missing + [""] * (max_len - len(missing)),
            "matched_count": [recon_summary.get("matched", 0)]
            + [""] * (max_len - 1),
        }
    )
    _write_formatted_xlsx(recon_df, recon_path)

    log.info("Reporting complete — 2 reports written to %s", OUTPUT_FOLDER)
    return billing_path, recon_path
